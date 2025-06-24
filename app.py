from flask import Flask, render_template, request, redirect, send_file, url_for
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, login_required, logout_user, UserMixin
import os
import io
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///history.db'
db = SQLAlchemy(app)

@app.route('/initdb')
def initdb():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        db.session.add(User(username='admin', password='1234'))
        db.session.commit()
    return 'База данных создана'
    
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(100))

class Record(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.String(100))
    user = db.Column(db.String(100))
    cur = db.Column(db.Float)
    new = db.Column(db.Float)
    cost = db.Column(db.Float)
    period = db.Column(db.Integer)
    payback = db.Column(db.String(100))
    economy = db.Column(db.Float)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = User.query.filter_by(username=request.form['username']).first()
        if u and u.password == request.form['password']:
            login_user(u)
            return redirect('/')
        return 'Неверные данные'
    return render_template('login.html')

@app.route('/logout')
def logout():
    logout_user()
    return redirect('/login')

@app.route('/')
@login_required
def index():
    records = Record.query.order_by(Record.id.desc()).all()
    return render_template('index.html', records=records)

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit(id):
    rec = Record.query.get_or_404(id)
    if request.method == 'POST':
        rec.user = request.form['user']
        rec.cur = float(request.form['cur'])
        rec.new = float(request.form['new'])
        rec.cost = float(request.form['cost'])
        rec.period = int(request.form['period'])
        rec.payback = request.form['payback']
        rec.economy = float(request.form['economy'])
        db.session.commit()
        return redirect('/')
    return render_template('edit.html', r=rec)

@app.route('/upload', methods=['POST'])
@login_required
def upload():
    file = request.files['excel']
    wb = load_workbook(file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = Record(
            date=row[0], user=row[1], cur=row[2], new=row[3], cost=row[4],
            period=row[5], payback=row[6], economy=row[7]
        )
        db.session.add(rec)
    db.session.commit()
    return redirect('/')

@app.route('/download')
@login_required
def download():
    wb = load_workbook('история.xlsx')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='история.xlsx')

if __name__ == '__main__':
    if not os.path.exists('history.db'):
        db.create_all()
        db.session.add(User(username='admin', password='1234'))
        db.session.commit()

    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
